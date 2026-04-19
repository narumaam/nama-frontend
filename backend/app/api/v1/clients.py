"""
Clients API — M13
POST /api/v1/clients/import          — bulk contact import (CSV/XLSX/Google/Outlook/vCard)
GET  /api/v1/clients/import/template — download CSV template
GET  /api/v1/clients/               — list clients (paginated, filterable)
POST /api/v1/clients/               — create single client
GET  /api/v1/clients/{id}           — get client
PATCH /api/v1/clients/{id}          — update client
"""
import io
import csv
import re
import logging
from datetime import datetime, timezone
from typing import List, Optional, Any, Dict

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.db.session import get_db
from app.api.v1.deps import require_tenant, can
from app.models.clients import Client, ClientStatus
from app.models.rbac import PermissionAuditLog

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Pydantic schemas ──────────────────────────────────────────────────────────

class ClientCreate(BaseModel):
    full_name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    secondary_phone: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = "India"
    status: Optional[str] = "ACTIVE"
    travel_type: Optional[str] = None
    preferred_destinations: Optional[List[str]] = []
    tags: Optional[List[str]] = []
    notes: Optional[str] = None
    total_bookings: Optional[int] = 0
    total_spend: Optional[float] = 0.0
    currency: Optional[str] = "INR"
    import_source: Optional[str] = None
    external_id: Optional[str] = None
    assigned_user_id: Optional[int] = None


class ClientUpdate(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    secondary_phone: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    status: Optional[str] = None
    travel_type: Optional[str] = None
    preferred_destinations: Optional[List[str]] = None
    tags: Optional[List[str]] = None
    notes: Optional[str] = None
    total_bookings: Optional[int] = None
    total_spend: Optional[float] = None
    currency: Optional[str] = None
    assigned_user_id: Optional[int] = None


class ClientOut(BaseModel):
    id: int
    tenant_id: int
    full_name: str
    email: Optional[str]
    phone: Optional[str]
    secondary_phone: Optional[str]
    city: Optional[str]
    country: Optional[str]
    status: str
    travel_type: Optional[str]
    preferred_destinations: Optional[List[str]]
    tags: Optional[List[str]]
    notes: Optional[str]
    total_bookings: int
    total_spend: float
    currency: str
    import_source: Optional[str]
    external_id: Optional[str]
    assigned_user_id: Optional[int]
    created_at: Optional[str]
    updated_at: Optional[str]
    last_contact: Optional[str]
    last_booking_date: Optional[str]

    class Config:
        from_attributes = True


def _client_to_dict(c: Client) -> dict:
    return {
        "id": c.id,
        "tenant_id": c.tenant_id,
        "full_name": c.full_name,
        "email": c.email,
        "phone": c.phone,
        "secondary_phone": c.secondary_phone,
        "city": c.city,
        "country": c.country,
        "status": c.status.value if hasattr(c.status, "value") else c.status,
        "travel_type": c.travel_type,
        "preferred_destinations": c.preferred_destinations or [],
        "tags": c.tags or [],
        "notes": c.notes,
        "total_bookings": c.total_bookings or 0,
        "total_spend": c.total_spend or 0.0,
        "currency": c.currency or "INR",
        "import_source": c.import_source,
        "external_id": c.external_id,
        "assigned_user_id": c.assigned_user_id,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        "last_contact": c.last_contact.isoformat() if c.last_contact else None,
        "last_booking_date": c.last_booking_date.isoformat() if c.last_booking_date else None,
    }


# ── Format detection ──────────────────────────────────────────────────────────

def _detect_format(headers: list) -> str:
    """
    Detect the source format of a CSV based on its header row.
    Returns: "google_contacts", "outlook", "vcf_csv", or "generic"
    """
    headers_lower = {h.strip().lower() for h in headers}

    # Google Contacts CSV
    google_signals = {"given name", "family name", "e-mail 1 - value"}
    if google_signals.issubset(headers_lower):
        return "google_contacts"

    # Outlook CSV
    outlook_signals = {"first name", "last name", "e-mail address"}
    if outlook_signals.issubset(headers_lower):
        return "outlook"

    return "generic"


# ── Column alias normalization ────────────────────────────────────────────────

_ALIAS_MAP: Dict[str, List[str]] = {
    "full_name":      ["name", "full name", "full_name", "contact name", "client name", "display name"],
    "email":          ["email", "email address", "e-mail", "e-mail 1 - value", "e-mail address",
                       "email 1 - value", "electronic mail address"],
    "phone":          ["phone", "phone number", "mobile", "mobile phone", "phone 1 - value",
                       "business phone", "whatsapp", "contact number", "cell", "tel",
                       "primary phone", "home phone"],
    "city":           ["city", "home city", "business city", "location", "town"],
    "country":        ["country", "home country", "business country", "nation", "country/region"],
    "tags":           ["tags", "labels", "groups", "category", "type", "group membership"],
    "travel_type":    ["travel type", "travel style", "segment", "category", "travel_type"],
    "notes":          ["notes", "comments", "remarks", "description", "note", "additional notes"],
    "total_spend":    ["total spend", "spend", "lifetime value", "ltv", "revenue", "total revenue",
                       "total_spend"],
    "total_bookings": ["total bookings", "bookings", "trips", "no of trips", "number of trips",
                       "total_bookings", "trip count"],
}


def _normalize_columns(df, detected_format: str):
    """
    Rename source DataFrame columns to standard NAMA field names.
    Handles Google Contacts / Outlook name concatenation.
    """
    import pandas as pd

    col_map = {}
    df_cols_lower = {c.strip().lower(): c for c in df.columns}

    # Handle Google Contacts name concatenation
    if detected_format == "google_contacts":
        given_col  = df_cols_lower.get("given name")
        family_col = df_cols_lower.get("family name")
        if given_col and family_col:
            df["full_name"] = (df[given_col].fillna("") + " " + df[family_col].fillna("")).str.strip()
            col_map[given_col]  = "__drop__"
            col_map[family_col] = "__drop__"

    # Handle Outlook name concatenation
    elif detected_format == "outlook":
        first_col = df_cols_lower.get("first name")
        last_col  = df_cols_lower.get("last name")
        if first_col and last_col:
            df["full_name"] = (df[first_col].fillna("") + " " + df[last_col].fillna("")).str.strip()
            col_map[first_col] = "__drop__"
            col_map[last_col]  = "__drop__"

    # Map remaining columns via alias table
    for target_field, aliases in _ALIAS_MAP.items():
        if target_field in df.columns:
            continue  # Already present (e.g. full_name from above)
        for alias in aliases:
            src_col = df_cols_lower.get(alias.lower())
            if src_col and src_col not in col_map:
                col_map[src_col] = target_field
                break

    # Apply renames, drop intermediate columns
    rename_map = {k: v for k, v in col_map.items() if v != "__drop__"}
    drop_cols  = [k for k, v in col_map.items() if v == "__drop__"]
    df = df.rename(columns=rename_map)
    df = df.drop(columns=[c for c in drop_cols if c in df.columns], errors="ignore")
    return df


# ── vCard parser ──────────────────────────────────────────────────────────────

def _parse_vcf(content: str) -> list:
    """Parse vCard (.vcf) file into list of contact dicts."""
    contacts = []
    current: dict = {}
    for line in content.splitlines():
        line = line.strip()
        if line == "BEGIN:VCARD":
            current = {}
        elif line == "END:VCARD":
            if current:
                contacts.append(current)
        elif line.startswith("FN:"):
            current["full_name"] = line[3:].strip()
        elif line.startswith("N:"):
            # N:Last;First;Middle;Prefix;Suffix
            parts = line[2:].split(";")
            last  = parts[0].strip() if len(parts) > 0 else ""
            first = parts[1].strip() if len(parts) > 1 else ""
            if not current.get("full_name"):
                current["full_name"] = f"{first} {last}".strip()
        elif "EMAIL" in line and ":" in line:
            current["email"] = line.split(":")[-1].strip()
        elif "TEL" in line and ":" in line:
            if "phone" not in current:
                current["phone"] = line.split(":")[-1].strip()
        elif line.startswith("ADR") and ":" in line:
            # ADR:;;Street;City;State;PostalCode;Country
            adr = line.split(":")[-1].split(";")
            if len(adr) > 3 and adr[3]:
                current["city"] = adr[3].strip()
            if len(adr) > 6 and adr[6]:
                current["country"] = adr[6].strip()
        elif line.startswith("NOTE:"):
            current["notes"] = line[5:].strip()
        elif line.startswith("ORG:"):
            org_note = f"Org: {line[4:].strip()}"
            existing = current.get("notes", "")
            current["notes"] = (f"{existing} | {org_note}" if existing else org_note).strip(" |")
    return contacts


# ── Helper: safe string → list ────────────────────────────────────────────────

def _parse_tags(value) -> list:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(t).strip() for t in value if str(t).strip()]
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return []
    # Split by comma or semicolon
    return [t.strip() for t in re.split(r"[,;]", s) if t.strip()]


def _safe_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in ("nan", "none", "") else s


def _safe_float(val) -> Optional[float]:
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _safe_int(val) -> Optional[int]:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return None


# ── Import endpoint ───────────────────────────────────────────────────────────

MAX_IMPORT_ROWS = 2000


@router.post("/import")
async def import_clients(
    file: UploadFile = File(...),
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    """
    Bulk import clients from CSV, XLSX, XLS, or VCF.
    Supports Google Contacts, Outlook, and generic CSV exports.
    Returns import summary: imported, skipped_duplicates, errors, total_rows, format_detected.
    """
    raw_bytes = await file.read()
    filename  = (file.filename or "").lower()

    imported           = 0
    skipped_duplicates = 0
    errors: list       = []
    format_detected    = "generic"
    rows: list         = []

    # ── Parse file by extension ───────────────────────────────────────────────
    try:
        if filename.endswith(".vcf"):
            format_detected = "vcf"
            content = raw_bytes.decode("utf-8", errors="replace")
            rows = _parse_vcf(content)

        elif filename.endswith((".xlsx", ".xls")):
            import pandas as pd
            df = pd.read_excel(io.BytesIO(raw_bytes), dtype=str)
            format_detected = _detect_format(list(df.columns))
            df = _normalize_columns(df, format_detected)
            rows = df.where(df.notna(), None).to_dict(orient="records")

        else:
            # CSV (default)
            import pandas as pd
            # Try to decode with UTF-8, fallback to latin-1
            try:
                text = raw_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text = raw_bytes.decode("latin-1")
            df = pd.read_csv(io.StringIO(text), dtype=str)
            format_detected = _detect_format(list(df.columns))
            df = _normalize_columns(df, format_detected)
            rows = df.where(df.notna(), None).to_dict(orient="records")

    except Exception as exc:
        logger.error("Client import parse error: %s", exc)
        raise HTTPException(status_code=400, detail=f"Could not parse file: {exc}")

    # Enforce row limit
    if len(rows) > MAX_IMPORT_ROWS:
        rows = rows[:MAX_IMPORT_ROWS]
        errors.append(f"File truncated to {MAX_IMPORT_ROWS} rows (limit).")

    total_rows = len(rows)

    # Pre-load existing emails + phones for dedup (within tenant)
    existing_emails = {
        r[0] for r in db.query(Client.email)
        .filter(Client.tenant_id == tenant_id, Client.email.isnot(None))
        .all()
    }
    existing_phones = {
        r[0] for r in db.query(Client.phone)
        .filter(Client.tenant_id == tenant_id, Client.phone.isnot(None))
        .all()
    }

    for idx, row in enumerate(rows):
        try:
            full_name = _safe_str(row.get("full_name")) or _safe_str(row.get("name"))
            email     = _safe_str(row.get("email"))
            phone     = _safe_str(row.get("phone"))

            # Skip rows with no identifying information
            if not full_name and not email and not phone:
                continue

            # Fallback full_name from email if missing
            if not full_name and email:
                full_name = email.split("@")[0].replace(".", " ").title()

            # Deduplication by email, then phone
            if email and email in existing_emails:
                skipped_duplicates += 1
                continue
            if phone and phone in existing_phones:
                skipped_duplicates += 1
                continue

            # Parse status
            raw_status = _safe_str(row.get("status")) or "ACTIVE"
            try:
                status = ClientStatus(raw_status.upper())
            except ValueError:
                status = ClientStatus.ACTIVE

            # Parse tags
            tags = _parse_tags(row.get("tags"))

            client = Client(
                tenant_id   = tenant_id,
                full_name   = full_name or "Unknown",
                email       = email,
                phone       = phone,
                secondary_phone = _safe_str(row.get("secondary_phone")),
                city        = _safe_str(row.get("city")),
                country     = _safe_str(row.get("country")) or "India",
                status      = status,
                travel_type = _safe_str(row.get("travel_type")),
                tags        = tags,
                notes       = _safe_str(row.get("notes")),
                total_spend    = _safe_float(row.get("total_spend")) or 0.0,
                total_bookings = _safe_int(row.get("total_bookings")) or 0,
                currency       = _safe_str(row.get("currency")) or "INR",
                import_source  = format_detected,
                preferred_destinations = [],
            )

            db.add(client)

            # Track in-memory for within-batch dedup
            if email:
                existing_emails.add(email)
            if phone:
                existing_phones.add(phone)

            imported += 1

        except Exception as row_exc:
            errors.append(f"Row {idx + 2}: {row_exc}")
            logger.warning("Client import row error at %d: %s", idx, row_exc)

    try:
        db.commit()
    except Exception as commit_exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"DB commit failed: {commit_exc}")

    return {
        "imported":           imported,
        "skipped_duplicates": skipped_duplicates,
        "errors":             errors[:20],   # cap error list for readability
        "total_rows":         total_rows,
        "format_detected":    format_detected,
    }


# ── Template download ─────────────────────────────────────────────────────────

@router.get("/import/template")
def clients_import_template():
    """Download a pre-filled CSV template for client import."""
    headers = [
        "full_name", "email", "phone", "city", "country",
        "travel_type", "tags", "total_spend", "total_bookings",
        "notes", "status",
    ]
    rows = [
        [
            "Rajesh Mehta", "rajesh.m@gmail.com", "+91 98765 43210",
            "Mumbai", "India", "Luxury", "VIP,Repeat", "1840000", "6",
            "Overwater bungalow fan", "VIP",
        ],
        [
            "Ananya Singh", "ananya.s@outlook.com", "+91 87654 32109",
            "Delhi", "India", "Family", "Family", "920000", "4",
            "Kids 8 and 12", "ACTIVE",
        ],
    ]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=nama_clients_template.csv"},
    )


# ── Export endpoint ───────────────────────────────────────────────────────────

@router.get("/export")
def export_clients(
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    _claims: dict = Depends(can("clients", "export")),
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    """
    Export clients as CSV or XLSX.
    Requires clients:export permission (enforced by can() dependency).
    Writes an audit log entry on every export.
    """
    # ── Build query ────────────────────────────────────────────────────────
    q = db.query(Client).filter(Client.tenant_id == tenant_id)
    if status:
        try:
            q = q.filter(Client.status == ClientStatus[status.upper()])
        except KeyError:
            pass
    if search:
        pattern = f"%{search}%"
        q = q.filter(
            (Client.full_name.ilike(pattern)) |
            (Client.email.ilike(pattern)) |
            (Client.phone.ilike(pattern)) |
            (Client.city.ilike(pattern))
        )
    clients = q.order_by(Client.full_name).all()

    # ── Audit log ──────────────────────────────────────────────────────────
    try:
        actor_id = int(_claims.get("user_id", 0)) if _claims else 0
        log_entry = PermissionAuditLog(
            tenant_id=tenant_id,
            actor_id=actor_id,
            action_type="export",
            module="clients",
            action_name="export",
            meta={
                "format": format,
                "filters": {"status": status, "search": search},
                "record_count": len(clients),
                "exported_at": datetime.now(timezone.utc).isoformat(),
            },
        )
        db.add(log_entry)
        db.commit()
    except Exception as e:
        logger.warning("export audit log failed: %s", e)

    # ── Column definitions ─────────────────────────────────────────────────
    HEADERS = [
        "ID", "Full Name", "Email", "Phone", "City", "Country",
        "Status", "Travel Type", "Tags", "Total Bookings",
        "Total Spend", "Currency", "Last Booking Date",
        "Preferred Destinations", "Notes", "Import Source", "Created At",
    ]

    def _row(c: Client) -> list:
        return [
            c.id,
            c.full_name,
            c.email or "",
            c.phone or "",
            c.city or "",
            c.country or "",
            c.status.value if c.status else "",
            c.travel_type or "",
            ", ".join(c.tags or []),
            c.total_bookings or 0,
            c.total_spend or 0,
            c.currency or "INR",
            c.last_booking_date.strftime("%Y-%m-%d") if c.last_booking_date else "",
            ", ".join(c.preferred_destinations or []),
            c.notes or "",
            c.import_source or "manual",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
        ]

    if format == "xlsx":
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        # Header row — bold + teal fill
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        # Data rows
        for row_idx, client in enumerate(clients, 2):
            for col_idx, value in enumerate(_row(client), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        # Auto-width (approximate)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"nama_clients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # Default: CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    for client in clients:
        writer.writerow(_row(client))
    output.seek(0)
    filename = f"nama_clients_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),  # utf-8-sig for Excel compatibility
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── CRUD endpoints ────────────────────────────────────────────────────────────

@router.get("/")
def list_clients(
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    q = db.query(Client).filter(Client.tenant_id == tenant_id)

    if status:
        try:
            q = q.filter(Client.status == ClientStatus(status.upper()))
        except ValueError:
            pass

    if search:
        term = f"%{search}%"
        q = q.filter(
            or_(
                Client.full_name.ilike(term),
                Client.email.ilike(term),
                Client.phone.ilike(term),
                Client.city.ilike(term),
            )
        )

    total = q.count()
    clients = q.order_by(Client.created_at.desc()).offset((page - 1) * size).limit(size).all()

    return {
        "clients": [_client_to_dict(c) for c in clients],
        "total":   total,
        "page":    page,
        "size":    size,
    }


@router.post("/")
def create_client(
    data: ClientCreate,
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    # Prevent duplicate email within tenant
    if data.email:
        existing = db.query(Client).filter(
            Client.tenant_id == tenant_id,
            Client.email == data.email,
        ).first()
        if existing:
            raise HTTPException(status_code=409, detail="A client with this email already exists.")

    try:
        status = ClientStatus(data.status.upper()) if data.status else ClientStatus.ACTIVE
    except ValueError:
        status = ClientStatus.ACTIVE

    client = Client(
        tenant_id            = tenant_id,
        full_name            = data.full_name,
        email                = data.email,
        phone                = data.phone,
        secondary_phone      = data.secondary_phone,
        city                 = data.city,
        country              = data.country or "India",
        status               = status,
        travel_type          = data.travel_type,
        preferred_destinations = data.preferred_destinations or [],
        tags                 = data.tags or [],
        notes                = data.notes,
        total_bookings       = data.total_bookings or 0,
        total_spend          = data.total_spend or 0.0,
        currency             = data.currency or "INR",
        import_source        = data.import_source or "manual",
        external_id          = data.external_id,
        assigned_user_id     = data.assigned_user_id,
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    return _client_to_dict(client)


@router.get("/{client_id}")
def get_client(
    client_id: int,
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    client = db.query(Client).filter(
        Client.id == client_id,
        Client.tenant_id == tenant_id,
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found.")
    return _client_to_dict(client)


@router.patch("/{client_id}")
def update_client(
    client_id: int,
    data: ClientUpdate,
    tenant_id: int = Depends(require_tenant),
    db: Session = Depends(get_db),
):
    client = db.query(Client).filter(
        Client.id == client_id,
        Client.tenant_id == tenant_id,
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found.")

    update_data = data.dict(exclude_unset=True)

    if "status" in update_data and update_data["status"]:
        try:
            update_data["status"] = ClientStatus(update_data["status"].upper())
        except ValueError:
            del update_data["status"]

    for field, value in update_data.items():
        setattr(client, field, value)

    client.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(client)
    return _client_to_dict(client)
